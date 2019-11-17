from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.hyperlink import Hyperlink
from flask import render_template, send_file
import os

## HTML INDEX VERSION
'''
def gen_file(index_rows):
    with open("file.html", "w") as f:
        #f = open('index_assiut.html', 'w')
        f.write(render_template('main_index.html', row_list=index_rows))
        #f.filename = 'indexmmm.html'
        #f.close()
        return f

'''

def index_init():
    file = load_workbook('app/templates/079254C-0000-ML-100_0.xlsx')
    sheet_list = file.sheetnames
    
    new_file = Workbook()
    new_ws = new_file.active

    index_rows=[]
    
    for sheet in sheet_list:
        sheet = file[sheet]
        sec1 = sheet['B13'].value
        sec2 = sheet['B15'].value
        sec3 = sheet['B17'].value
        if sheet['B17'].value is not None:
                para2 = sec3.split(' ')[0]
        else:
                para2 = sec2.split(' ')[0]
        for row in sheet.iter_rows(min_row=2):
            cell = row[2]
            if cell.value is not None and cell.value[0:7] == '079254C':
                
                cell_start_value = str(cell.value)
                cell.hyperlink = Hyperlink('link',
                        target='.\\FILE\\'+ para2 +'\\' + str(cell_start_value) + "_" + str(row[3].value),
                        display='something')

                print(cell.value)

                link1 = row[3]
                link1.hyperlink = Hyperlink('link',
                        target='.\\FILE\\'+ para2 +'\\' + cell_start_value + "_" + str(row[3].value),
                        display='something')
                
                # Rows for the html version 
                index_rows.append((cell_start_value,cell.value))

                # Add Row to the Client Report
                try:
                        new_ws.append((sec1,sec2,sec3,row[1].value,cell_start_value,row[3].value, row[4].value))  
                except:
                        new_ws.append((sec1,sec2,sec3,row[1].value,cell_start_value,row[3].value))  

    file.save('tavole.xlsx')
    new_file.save('report_cliente.xlsx')
    '''
    f = gen_file(index_rows)
    
    return send_file(f, as_attachment=True, attachment_filename='indexmain.html')
    '''
    #
    # Generate folder and file list
    #

    new_file = Workbook()
    listafolder = new_file.active
    for path, folder, file in os.walk('./files'):
        if isinstance(folder, list):
            for f in folder:
                listafolder.append([f])
        else:
            listafolder.append([folder])
    
    new_file.save('listafolder.xlsx')
    
    new_file = Workbook()
    listafile = new_file.active
    for path, folder, file in os.walk('./files'):
        if isinstance(file, list):
            for f in file:
                listafile.append([path,f.split('.')[0],f.split('.')[1]])
            else:
                listafile.append([path,file.split('.')[0],file.split('.')[1]])
    
    new_file.save('listafile.xlsx')
    print('-------- ------------ -------- -------- ------ ------ --- -')
    print('           -------- ----- INDEX DONE -------- -------')
    print('-------- ------------ -------- -------- ------ ------ --- -')

#index_init()   
 
def dict_doc():
        wb = load_workbook('testof1.xlsx')
        ws = wb.active

        #for row in ws.iter_rows:





def filelist():
        new_file = Workbook()
        listafolder = new_file.active
        for path, folder, file in os.walk('./files'):
                if isinstance(folder, list):
                        for f in folder:
                                listafolder.append([f])
                else:
                        listafolder.append([folder])
        new_file.save('listafolder.xlsx')
        
        new_file = Workbook()
        listafile = new_file.active
        for path, folder, file in os.walk('./files'):
                if isinstance(file, list):
                        for f in file:

                                listafile.append([path,f.split('.')[0],f.split('.')[1]])
                else:
                        listafile.append([path,file.split('.')[0],file.split('.')[1]])
        new_file.save('listafile.xlsx') 
#filelist()         